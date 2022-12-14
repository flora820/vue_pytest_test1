from pathlib import Path
from openpyxl import load_workbook
from common.exception_utils import *
import time

@exception_utils
class ExcelUtil(object):

    def __init__(self, excel_path='%s/data/case_excel/接口测试框架实践用例.xlsx' % Path(__file__).parent.parent):
        self.wb = load_workbook(excel_path)
        self.base_dir = Path(__file__).parent.parent
        self.template = '{"id":0, "url":"","case_name":"","method":"","header":"","body":"",\
                "expect_result":"","actual_result":"","valiadate":"","smoke":""},'

    @exception_utils
    def read_excel(self):
        value = []
        smoke_value = []

        for sheetname in self.wb.sheetnames:
            ws = self.wb[sheetname]
            case_num = len(list(ws.values)) - 1  # 一个sheet中用例的数量
            case_list = list(ws.values)
            case_list.pop(0)  #去除表头
            case_template = self.template * case_num
            case_template_list = eval("[" + case_template[:-1]+ "]")


            for i in range(len(case_list)): #i 第i个用例
                case_template_list[i]['id']        = case_list[i][0]
                case_template_list[i]['url']       = case_list[i][1]
                case_template_list[i]['case_name'] = case_list[i][2]
                case_template_list[i]['method']    = case_list[i][3]
                case_template_list[i]['header']    = case_list[i][4]
                case_template_list[i]['body']        = case_list[i][5]
                case_template_list[i]['expect_result']        = case_list[i][6]
                case_template_list[i]['actual_result']        = case_list[i][7]
                case_template_list[i]['valiadate']        = case_list[i][8]
                case_template_list[i]['smoke']        = case_list[i][9]
         #   value.append({"cases":case_template_list})
            
            value.append({str(sheetname):case_template_list})
       # print("value:\t"+str(value)+"\n")
        
        #for v in value:
        for v in value:
            for sheetcase,total_case in v.items():
                smoke_case=[]
                print(len(total_case))
                for case in total_case:
                    if 'yes' in str(case['smoke']):
                        smoke_case.append(case)
                smoke_value.append({sheetcase:smoke_case})

       # smoke={"smoke":smoke_value}
       # print("smoke:\t"+str(smoke)+"\n")
        #print("smoke:\t"+str(smoke_value)+"\n")

        return value,smoke_value


if __name__ == '__main__':
    excel_path='%s/data/excel/vue_testcase.xlsx' % base_dir
    ExcelUtil(excel_path).read_excel()
